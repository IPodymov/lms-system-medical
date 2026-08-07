from io import BytesIO

from django.contrib import messages
from django.contrib.auth import login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.password_validation import validate_password
from django.contrib.auth.views import LoginView
from django.core import signing
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.paginator import Paginator
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import Resolver404, resolve, reverse
from openpyxl import Workbook, load_workbook

from apps.courses.models import CourseEnrollmentLink, CourseRunStaff
from apps.learning.models import Enrollment
from apps.learning.selectors import with_block_counts
from apps.organizations.models import (
    OrganizationMembership,
    StudyGroup,
    StudyGroupMember,
)

from . import services
from .forms import (
    CourseEnrollmentForm,
    CourseStaffForm,
    EnrollmentLinkForm,
    GroupStudentForm,
    ManagedUserForm,
    OrganizationForm,
    ProfileForm,
    RegistrationForm,
    StudentImportForm,
    StudyGroupForm,
    UserPasswordForm,
)
from .models import User
from .selectors import can_access_documentation as _can_access_documentation
from .selectors import can_access_management_documentation as _can_access_management_documentation
from .selectors import can_manage_users as _can_manage_users
from .selectors import managed_course_runs as _managed_course_runs
from .selectors import managed_memberships as _managed_memberships
from .selectors import managed_organizations as _managed_organizations
from .selectors import managed_study_groups as _managed_study_groups
from .selectors import request_organization as _request_organization
from .services import add_student_to_group as _add_student_to_group
from .services import default_department as _default_department
from .services import get_or_create_user as _get_or_create_user
from .services import organization_slug as _organization_slug

ADMIN_DOCUMENTATION_SALT = "medical-lms.admin-documentation"
ADMIN_DOCUMENTATION_MAX_AGE = 60 * 60 * 24
ENROLLMENT_LINKS_PER_PAGE = 20
# Дашборд показывает не весь список, а несколько курсов «продолжить с того
# места, где остановились»; полный список живёт на странице «Мои курсы».
DASHBOARD_ENROLLMENTS = 6


def _admin_documentation_url(request):
    token = signing.dumps(str(request.user.pk), salt=ADMIN_DOCUMENTATION_SALT, compress=True)
    return reverse("admin-documentation", args=[token])


class Login(LoginView):
    template_name = "accounts/login.html"


def register(request):
    if request.user.is_authenticated:
        return redirect("dashboard")
    form = RegistrationForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = form.save()
        login(request, user)
        messages.success(request, "Аккаунт создан. Добро пожаловать в МедЛМС.")
        return redirect("dashboard")
    return render(request, "accounts/register.html", {"form": form})


def logout_view(request):
    logout(request)
    return redirect("login")


@login_required
def dashboard(request):
    enrollments = with_block_counts(
        Enrollment.objects.filter(user=request.user, status="active").select_related(
            "course_run__course"
        )
    )
    return render(
        request,
        "dashboard.html",
        {"enrollments": enrollments[:DASHBOARD_ENROLLMENTS]},
    )


@login_required
def profile(request):
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "profile":
            profile_form = ProfileForm(request.POST, request.FILES, instance=request.user)
            password_form = UserPasswordForm(request.user)
            if profile_form.is_valid():
                profile_form.save()
                messages.success(request, "Профиль сохранён.")
                return redirect("profile")
        elif action == "password":
            profile_form = ProfileForm(instance=request.user)
            password_form = UserPasswordForm(request.user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, "Пароль изменён.")
                return redirect("profile")
        else:
            profile_form = ProfileForm(instance=request.user)
            password_form = UserPasswordForm(request.user)
    else:
        profile_form = ProfileForm(instance=request.user)
        password_form = UserPasswordForm(request.user)
    # Счётчики блоков нужны общей карточке записи: без них она показывает
    # «Материалы готовятся» вместо «Пройдено 7 из 16».
    history = with_block_counts(
        request.user.enrollments.select_related("course_run__course")
    ).order_by("-updated_at")
    return render(
        request,
        "accounts/profile.html",
        {
            "profile_form": profile_form,
            "password_form": password_form,
            "history": history,
        },
    )


def _validate_password_or_message(request, password) -> bool:
    try:
        validate_password(password)
    except ValidationError as error:
        messages.error(request, " ".join(error.messages))
        return False
    return True


def _study_group_id_from_url(url: str) -> int | None:
    """Достать id группы из адреса возврата, если он ведёт на страницу группы."""
    try:
        match = resolve(url)
    except (Resolver404, ValueError):
        return None
    return match.kwargs.get("group_id") if match.url_name == "study-group-detail" else None


def _dashboard_forms(request, managed_organizations, course_runs, **bound):
    """Собрать семь форм страницы управления.

    Именованный аргумент подменяет форму связанной: view, поймавший
    невалидный POST, отдаёт страницу целиком со своей формой, а остальные
    шесть остаются пустыми. Иначе показать ошибку у поля нельзя — при
    редиректе форма и введённые значения теряются.
    """
    teacher_users = User.objects.filter(
        memberships__organization__in=managed_organizations,
        memberships__role=OrganizationMembership.Role.TEACHER,
        memberships__status="active",
    ).order_by("last_name", "first_name", "email")
    student_users = User.objects.filter(
        memberships__organization__in=managed_organizations,
        memberships__role=OrganizationMembership.Role.STUDENT,
        memberships__status="active",
    ).order_by("last_name", "first_name", "email")
    forms_by_name = {
        "organization_form": OrganizationForm(),
        "study_group_form": StudyGroupForm(organizations=managed_organizations),
        "import_form": StudentImportForm(organizations=managed_organizations),
        "user_form": ManagedUserForm(
            organizations=managed_organizations,
            roles=OrganizationMembership.Role.choices,
            initial={"role": OrganizationMembership.Role.STUDENT},
        ),
        "staff_form": CourseStaffForm(course_runs=course_runs, teachers=teacher_users),
        "enrollment_form": CourseEnrollmentForm(course_runs=course_runs, students=student_users),
        "link_form": EnrollmentLinkForm(course_runs=course_runs),
    }
    forms_by_name.update(bound)
    return forms_by_name


def _dashboard_context(request, **bound_forms):
    memberships = _managed_memberships(request.user)
    managed_organizations = _managed_organizations(request.user)
    if not managed_organizations.exists():
        if not request.user.is_superuser:
            raise PermissionDenied
        return {
            "metrics": {"users": 0, "students": 0, "completed": 0},
            "can_manage_users": False,
            "can_create_organizations": True,
            "organizations": [],
            "course_runs": [],
            "study_groups": [],
            "admin_documentation_url": _admin_documentation_url(request),
            "organization_form": bound_forms.get("organization_form") or OrganizationForm(),
        }
    if request.user.is_superuser and services.bootstrap_superuser_membership(
        request.user, managed_organizations
    ):
        memberships = _managed_memberships(request.user)
    enrollments = Enrollment.objects.filter(
        course_run__course__organization__in=managed_organizations
    )
    if not request.user.is_superuser:
        run_ids = CourseRunStaff.objects.filter(
            user=request.user, role__in=["teacher", "assistant"]
        ).values("course_run_id")
        enrollments = enrollments.filter(course_run_id__in=run_ids)
    course_runs = _managed_course_runs(request.user).order_by("course__title", "title")
    study_groups = _managed_study_groups(request.user).order_by("name", "admission_year")
    enrollment_links = Paginator(
        CourseEnrollmentLink.objects.filter(course_run__in=course_runs)
        .select_related("course_run__course")
        .order_by("-created_at"),
        ENROLLMENT_LINKS_PER_PAGE,
    ).get_page(request.GET.get("page"))
    context = {
        "metrics": {
            "users": memberships.values("user_id").distinct().count(),
            "students": enrollments.values("user_id").distinct().count(),
            "completed": enrollments.filter(status="completed").count(),
        },
        "can_manage_users": _can_manage_users(request.user),
        "can_create_organizations": request.user.is_superuser,
        "organizations": managed_organizations,
        "course_runs": course_runs,
        "study_groups": study_groups,
        "enrollment_links": enrollment_links,
        "admin_documentation_url": (
            _admin_documentation_url(request) if request.user.is_superuser else None
        ),
    }
    context.update(_dashboard_forms(request, managed_organizations, course_runs, **bound_forms))
    return context


def _dashboard_with_errors(request, **bound_forms):
    """Отдать страницу управления с невалидной формой (HTTP 200, не редирект)."""
    return render(
        request, "accounts/admin_dashboard.html", _dashboard_context(request, **bound_forms)
    )


@login_required
def admin_dashboard(request):
    return render(request, "accounts/admin_dashboard.html", _dashboard_context(request))


@login_required
def admin_documentation(request, token):
    if not request.user.is_superuser:
        raise PermissionDenied
    try:
        user_id = signing.loads(
            token,
            salt=ADMIN_DOCUMENTATION_SALT,
            max_age=ADMIN_DOCUMENTATION_MAX_AGE,
        )
    except signing.BadSignature as error:
        raise PermissionDenied from error
    if user_id != str(request.user.pk):
        raise PermissionDenied
    return render(request, "accounts/admin_documentation.html")


@login_required
def documentation_home(request):
    if not _can_access_documentation(request.user):
        raise PermissionDenied
    return render(
        request,
        "accounts/documentation_home.html",
        {"can_access_management_documentation": _can_access_management_documentation(request.user)},
    )


@login_required
def course_documentation(request):
    if not _can_access_documentation(request.user):
        raise PermissionDenied
    return render(
        request,
        "accounts/documentation_courses.html",
        {"can_access_management_documentation": _can_access_management_documentation(request.user)},
    )


@login_required
def management_documentation(request):
    if not _can_access_management_documentation(request.user):
        raise PermissionDenied
    return render(
        request,
        "accounts/documentation_management.html",
        {"can_access_management_documentation": True},
    )


@login_required
def add_organization(request):
    if request.method != "POST" or not request.user.is_superuser:
        raise PermissionDenied

    form = OrganizationForm(request.POST)
    if not form.is_valid():
        return _dashboard_with_errors(request, organization_form=form)

    organization = form.save(commit=False)
    organization.slug = _organization_slug(organization.name)
    organization.save()
    OrganizationMembership.objects.update_or_create(
        user=request.user,
        organization=organization,
        defaults={"role": OrganizationMembership.Role.SYSTEM_ADMIN, "status": "active"},
    )
    messages.success(request, "Организация создана. Теперь можно создавать и редактировать курсы.")
    return redirect("admin-dashboard")


@login_required
def manage_user_role(request, membership_id):
    if request.method != "POST" or not _can_manage_users(request.user):
        raise PermissionDenied
    membership = get_object_or_404(_managed_memberships(request.user), pk=membership_id)
    role = request.POST.get("role")
    if role in OrganizationMembership.Role.values:
        membership.role = role
        membership.save(update_fields=["role", "updated_at"])
        messages.success(request, "Роль пользователя обновлена.")
    return redirect("admin-dashboard")


@login_required
def add_user(request):
    if request.method != "POST" or not _can_manage_users(request.user):
        raise PermissionDenied
    managed_organizations = _managed_organizations(request.user)
    form = ManagedUserForm(
        request.POST,
        organizations=managed_organizations,
        roles=OrganizationMembership.Role.choices,
    )
    if not form.is_valid():
        return _dashboard_with_errors(request, user_form=form)

    data = form.cleaned_data
    user, created = User.objects.get_or_create(
        email=data["email"], defaults={"username": data["username"]}
    )
    if created:
        user.set_password(data["password"])
        user.first_name = data["first_name"]
        user.last_name = data["last_name"]
        user.middle_name = data["middle_name"]
        user.save(update_fields=["password", "first_name", "last_name", "middle_name"])
    OrganizationMembership.objects.update_or_create(
        user=user,
        organization=data["organization"],
        defaults={"role": data["role"], "status": "active"},
    )
    messages.success(
        request,
        "Пользователь добавлен. Передайте ему заданный временный пароль безопасным способом.",
    )
    return redirect("admin-dashboard")


@login_required
def add_teacher(request):
    """Only a system administrator can add teachers across the college."""
    if request.method != "POST" or not request.user.is_superuser:
        raise PermissionDenied
    organization = _request_organization(request.user, request.POST.get("organization_id"))
    email = request.POST.get("email", "").strip().lower()
    password = request.POST.get("password", "")
    if not organization or not email or not password:
        messages.error(request, "Укажите организацию, email и временный пароль преподавателя.")
        return redirect("admin-dashboard")
    if not _validate_password_or_message(request, password):
        return redirect("admin-dashboard")
    with transaction.atomic():
        user, _ = _get_or_create_user(
            email=email,
            password=password,
            first_name=request.POST.get("first_name", "").strip(),
            last_name=request.POST.get("last_name", "").strip(),
            middle_name=request.POST.get("middle_name", "").strip(),
        )
        OrganizationMembership.objects.update_or_create(
            user=user,
            organization=organization,
            defaults={
                "role": OrganizationMembership.Role.TEACHER,
                "status": "active",
                "employee_number": request.POST.get("employee_number", "").strip(),
            },
        )
    messages.success(request, "Преподаватель добавлен в систему.")
    return redirect("admin-dashboard")


@login_required
def add_study_group(request):
    if request.method != "POST" or not _can_manage_users(request.user):
        raise PermissionDenied
    form = StudyGroupForm(request.POST, organizations=_managed_organizations(request.user))
    if not form.is_valid():
        return _dashboard_with_errors(request, study_group_form=form)

    data = form.cleaned_data
    group, created = StudyGroup.objects.get_or_create(
        department=_default_department(data["organization"]),
        name=data["name"],
        admission_year=data["admission_year"],
        defaults={"graduation_year": data["graduation_year"]},
    )
    if not created:
        group.graduation_year = data["graduation_year"]
        group.save(update_fields=["graduation_year", "updated_at"])
    messages.success(request, f"Учебная группа «{data['name']}» сохранена.")
    return redirect("admin-dashboard")


def _group_context(**context):
    """Контекст страницы группы: формы по умолчанию плюс переданные связанные."""
    context.setdefault("student_form", GroupStudentForm())
    context.setdefault("group_link_placeholder", f"Для {context['study_group'].name}")
    context.setdefault("link_form", EnrollmentLinkForm(course_runs=context["course_runs"]))
    return context


@login_required
def study_group_detail(request, group_id, **group_forms):
    if not _can_manage_users(request.user):
        raise PermissionDenied
    study_group = get_object_or_404(_managed_study_groups(request.user), pk=group_id)
    organization = study_group.department.faculty.organization
    members = (
        StudyGroupMember.objects.filter(study_group=study_group, left_at__isnull=True)
        .select_related("user")
        .order_by("user__last_name", "user__first_name", "user__email")
    )
    members = list(members)
    memberships_by_user_id = {
        membership.user_id: membership
        for membership in OrganizationMembership.objects.filter(
            organization=organization,
            user_id__in=[member.user_id for member in members],
        )
    }
    for member in members:
        member.organization_membership = memberships_by_user_id.get(member.user_id)
    course_runs = _managed_course_runs(request.user).filter(course__organization=organization)
    enrollment_links = Paginator(
        CourseEnrollmentLink.objects.filter(course_run__in=course_runs)
        .select_related("course_run__course")
        .order_by("-created_at"),
        ENROLLMENT_LINKS_PER_PAGE,
    ).get_page(request.GET.get("page"))
    return render(
        request,
        "accounts/study_group_detail.html",
        _group_context(
            study_group=study_group,
            members=members,
            course_runs=course_runs.order_by("course__title", "title"),
            enrollment_links=enrollment_links,
            **group_forms,
        ),
    )


@login_required
def add_student(request):
    if request.method != "POST" or not _can_manage_users(request.user):
        raise PermissionDenied
    study_group = get_object_or_404(
        _managed_study_groups(request.user), pk=request.POST.get("group_id")
    )
    organization = study_group.department.faculty.organization
    form = GroupStudentForm(request.POST)
    if not form.is_valid():
        # Страница группы отдаётся целиком: у формы шесть полей, и сказать
        # «что-то не так» поверх редиректа значило бы потерять введённое.
        return study_group_detail(request, study_group.pk, student_form=form)

    data = form.cleaned_data
    with transaction.atomic():
        user, _ = _get_or_create_user(
            email=data["email"],
            password=data["password"],
            first_name=data["first_name"],
            last_name=data["last_name"],
            middle_name=data["middle_name"],
        )
        _add_student_to_group(
            user=user,
            organization=organization,
            study_group=study_group,
            student_number=data["student_number"],
        )
    messages.success(request, "Студент создан и добавлен в учебную группу.")
    return redirect("admin-dashboard")


@login_required
def import_students(request):
    if request.method != "POST" or not _can_manage_users(request.user):
        raise PermissionDenied
    form = StudentImportForm(
        request.POST, request.FILES, organizations=_managed_organizations(request.user)
    )
    if not form.is_valid():
        return _dashboard_with_errors(request, import_form=form)
    organization = form.cleaned_data["organization"]
    spreadsheet = form.cleaned_data["spreadsheet"]
    try:
        workbook = load_workbook(spreadsheet, read_only=True, data_only=True)
        parsed_rows = services.parse_student_import_rows(workbook)
    except (OSError, ValueError) as error:
        messages.error(request, f"Не удалось импортировать файл: {error}")
        return redirect("admin-dashboard")

    created_students, generated_credentials = services.import_students(
        organization=organization, parsed_rows=parsed_rows
    )
    if generated_credentials:
        result = Workbook()
        worksheet = result.active
        worksheet.title = "Учётные записи"
        worksheet.append(
            ["Фамилия", "Имя", "Группа", "Username", "Логин (email)", "Временный пароль"]
        )
        for credential in generated_credentials:
            worksheet.append(credential)
        content = BytesIO()
        result.save(content)
        response = HttpResponse(
            content.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="student_credentials.xlsx"'
        return response
    messages.success(
        request,
        "Импорт завершён: обработано "
        f"{len(parsed_rows)} строк, создано студентов: {created_students}.",
    )
    return redirect("admin-dashboard")


@login_required
def assign_course_staff(request):
    if request.method != "POST" or not _can_manage_users(request.user):
        raise PermissionDenied
    course_runs = _managed_course_runs(request.user)
    teacher_users = User.objects.filter(
        memberships__organization__in=_managed_organizations(request.user),
        memberships__role=OrganizationMembership.Role.TEACHER,
        memberships__status="active",
    )
    form = CourseStaffForm(request.POST, course_runs=course_runs, teachers=teacher_users)
    if not form.is_valid():
        return _dashboard_with_errors(request, staff_form=form)

    data = form.cleaned_data
    CourseRunStaff.objects.update_or_create(
        course_run=data["course_run"], user=data["user"], defaults={"role": data["role"]}
    )
    messages.success(request, "Состав преподавателей и кураторов курса обновлён.")
    return redirect("admin-dashboard")


@login_required
def remove_course_staff(request, staff_id):
    if request.method != "POST" or not _can_manage_users(request.user):
        raise PermissionDenied
    staff = get_object_or_404(
        CourseRunStaff.objects.filter(course_run__in=_managed_course_runs(request.user)),
        pk=staff_id,
    )
    staff.delete()
    messages.success(request, "Сотрудник удалён из состава курса.")
    return redirect("admin-dashboard")


@login_required
def manage_course_enrollment(request):
    if request.method != "POST" or not _can_manage_users(request.user):
        raise PermissionDenied
    action = request.POST.get("action")
    if action == "remove":
        enrollment = get_object_or_404(
            Enrollment.objects.filter(course_run__in=_managed_course_runs(request.user)),
            pk=request.POST.get("enrollment_id"),
        )
        enrollment.delete()
        messages.success(request, "Студент исключён из курса.")
        return redirect("admin-dashboard")

    if action == "add_student":
        # Поток и слушателя проверяет форма: она же покажет ошибку у нужного
        # поля, если выбранный слушатель не из организации курса.
        student_users = User.objects.filter(
            memberships__organization__in=_managed_organizations(request.user),
            memberships__role=OrganizationMembership.Role.STUDENT,
            memberships__status="active",
        )
        form = CourseEnrollmentForm(
            request.POST,
            course_runs=_managed_course_runs(request.user),
            students=student_users,
        )
        if not form.is_valid():
            return _dashboard_with_errors(request, enrollment_form=form)
        Enrollment.objects.get_or_create(
            course_run=form.cleaned_data["course_run"],
            user=form.cleaned_data["user"],
            defaults={"status": Enrollment.Status.ACTIVE, "enrollment_source": "manual"},
        )
        messages.success(request, "Студент добавлен в курс.")
        return redirect("admin-dashboard")

    # Массовая запись приходит со страницы группы, где поток и группа
    # выбираются из списков в контексте самой группы.
    course_run = get_object_or_404(
        _managed_course_runs(request.user), pk=request.POST.get("course_run_id")
    )
    if action == "add_group":
        group = get_object_or_404(
            _managed_study_groups(request.user), pk=request.POST.get("group_id")
        )
        if group.department.faculty.organization_id != course_run.course.organization_id:
            raise PermissionDenied
        user_ids = group.studygroupmember_set.filter(left_at__isnull=True).values_list(
            "user_id", flat=True
        )
        Enrollment.objects.bulk_create(
            [
                Enrollment(
                    course_run=course_run,
                    user_id=user_id,
                    status=Enrollment.Status.ACTIVE,
                    enrollment_source="group",
                )
                for user_id in user_ids
            ],
            ignore_conflicts=True,
        )
        messages.success(request, "Активные студенты группы добавлены в курс.")
    else:
        raise PermissionDenied
    return redirect("admin-dashboard")


@login_required
def create_course_enrollment_link(request):
    if request.method != "POST" or not _can_manage_users(request.user):
        raise PermissionDenied
    form = EnrollmentLinkForm(request.POST, course_runs=_managed_course_runs(request.user))
    if not form.is_valid():
        # Форма живёт на двух страницах, поэтому возвращаемся на ту, откуда
        # пришли: next уже несёт этот адрес.
        group_id = _study_group_id_from_url(request.POST.get("next", ""))
        if group_id:
            return study_group_detail(request, group_id, link_form=form)
        return _dashboard_with_errors(request, link_form=form)

    data = form.cleaned_data
    CourseEnrollmentLink.objects.create(
        course_run=data["course_run"],
        created_by=request.user,
        label=data["label"],
        expires_at=data["expires_at"],
    )
    messages.success(request, "Ссылка для записи на курс создана.")
    return redirect(request.POST.get("next") or "admin-dashboard")


@login_required
def deactivate_course_enrollment_link(request, link_id):
    if request.method != "POST" or not _can_manage_users(request.user):
        raise PermissionDenied
    enrollment_link = get_object_or_404(
        CourseEnrollmentLink.objects.filter(course_run__in=_managed_course_runs(request.user)),
        pk=link_id,
    )
    enrollment_link.is_active = False
    enrollment_link.save(update_fields=["is_active", "updated_at"])
    messages.success(request, "Ссылка для записи отключена.")
    return redirect(request.POST.get("next") or "admin-dashboard")
